"""
Mecklenburg County Polaris3G Tax Bill Scraper
Reads parcel IDs from Excel column P (rows 3-227), scrapes 2024 tax bill data,
and writes results back to columns J, K, L, M.

SETUP (run once):
    pip install playwright openpyxl
    playwright install chromium

USAGE:
    python scrape_tax_bills.py --file "path/to/your/file.xlsx"
"""

import asyncio
import argparse
import csv
import sys
import time
import re
from pathlib import Path

import openpyxl
from playwright.async_api import async_playwright, TimeoutError as PWTimeout

# ── Configuration ─────────────────────────────────────────────────────────────
BASE_URL = "https://polaris3g.mecklenburgcountync.gov/"
FIRST_DATA_ROW = 3
LAST_DATA_ROW = 227
COL_PARCEL   = 16   # P
COL_ASSESSED = 10   # J
COL_MILLAGE  = 11   # K
COL_DIRECT   = 12   # L
COL_INTEREST = 13   # M
COL_FORMULA  = 14   # N  (already has formula — read-only for validation)
SAVE_EVERY   = 10   # save workbook every N parcels
DELAY_SECS   = 1.5  # pause between parcels

# ── Helpers ───────────────────────────────────────────────────────────────────

def parse_money(text: str) -> float:
    """Convert '$1,234.56' or '1234.56' to float, return 0.0 on failure."""
    if not text:
        return 0.0
    cleaned = re.sub(r"[^\d.\-]", "", text.strip())
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def col_letter(n: int) -> str:
    """Column index (1-based) → letter, e.g. 10 → J."""
    result = ""
    while n:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result


# ── Core scraper ──────────────────────────────────────────────────────────────

async def scrape_parcel(page, parcel_id: str) -> dict | None:
    """
    Navigate Polaris3G for one parcel and return scraped values, or None on failure.
    Returns dict with keys: assessed, millage, direct_assessments, interest, total_tax_bill
    """
    # ── 1. Search ──────────────────────────────────────────────────────────────
    await page.goto(BASE_URL, wait_until="networkidle", timeout=30_000)

    # The search input is usually a text field near the top of the page
    search_sel = "input[placeholder*='Search'], input[placeholder*='search'], input[type='search'], #searchInput, .search-input"
    try:
        await page.wait_for_selector(search_sel, timeout=10_000)
    except PWTimeout:
        # Fall back: look for any visible input in the header area
        search_sel = "input"

    search_box = page.locator(search_sel).first
    await search_box.click()
    await search_box.fill("")
    await search_box.type(parcel_id, delay=50)
    await search_box.press("Enter")

    # Wait for results panel to update
    await page.wait_for_timeout(2_000)

    # Check for "no results" indicators
    no_results_texts = ["no results", "not found", "0 results"]
    body_text = (await page.inner_text("body")).lower()
    if any(t in body_text for t in no_results_texts):
        return None

    # ── 2. Click the first result (parcel link) ────────────────────────────────
    result_sel = ".search-result, .result-item, [class*='result'] a, .parcel-result"
    try:
        await page.wait_for_selector(result_sel, timeout=8_000)
        await page.locator(result_sel).first.click()
        await page.wait_for_timeout(1_500)
    except PWTimeout:
        # Maybe clicking Enter already navigated directly
        pass

    # ── 3. Find "Tax Information" section in left panel ───────────────────────
    tax_info_sel = "text=Tax Information, text=Tax Bill Information, [class*='tax']"
    try:
        await page.wait_for_selector(tax_info_sel, timeout=10_000)
    except PWTimeout:
        return None

    # Scroll it into view and click
    tax_link = page.get_by_text("Tax Bill Information", exact=False).first
    try:
        await tax_link.scroll_into_view_if_needed()
        await tax_link.click()
        await page.wait_for_timeout(2_000)
    except Exception:
        return None

    # ── 4. Select 2024 tax bill ───────────────────────────────────────────────
    # There may be a year selector — prefer "2024"
    year_2024 = page.get_by_text("2024", exact=False)
    count = await year_2024.count()
    if count > 0:
        await year_2024.first.click()
        await page.wait_for_timeout(1_500)

    # Confirm we're on 2024; if first option is 2025, try clicking the second
    current_content = (await page.inner_text("body")).lower()
    if "2025" in current_content and "2024" not in current_content:
        # Try explicit "2024" radio or select option
        try:
            await page.select_option("select", label="2024")
            await page.wait_for_timeout(1_500)
        except Exception:
            pass

    # ── 5. Scrape the values ──────────────────────────────────────────────────
    full_text = await page.inner_text("body")

    # Assessed Value
    assessed = _extract_after(full_text, ["assessed value", "total assessed value"])

    # Millage Rate — look for "combined rate" or "total rate"
    millage = _extract_after(full_text, ["combined rate", "total millage", "total rate", "millage rate"])
    # Millage is usually expressed as X.XXXX (per $100), store as decimal
    # e.g. "0.7572" stays as-is; "75.72" per $1000 → divide by 10
    if millage > 10:
        millage = millage / 1000  # some portals show per $1000

    # Direct Assessments — sum all line items under "Direct Assessments"
    direct_total = _extract_direct_assessments(full_text)

    # Interest Due
    interest = _extract_after(full_text, ["interest due", "interest", "penalty & interest"])

    # Total Tax Bill (for validation)
    total_bill = _extract_after(full_text, ["total tax bill", "total bill", "amount due", "total due"])

    if assessed == 0.0 and total_bill == 0.0:
        # Didn't get meaningful data — probably on wrong page
        return None

    return {
        "assessed": assessed,
        "millage": millage,
        "direct_assessments": direct_total,
        "interest": interest,
        "total_tax_bill": total_bill,
    }


def _extract_after(text: str, labels: list[str]) -> float:
    """
    Find the first occurrence of any label and extract the dollar amount that
    follows it on the same line or the next non-empty line.
    """
    lower = text.lower()
    for label in labels:
        idx = lower.find(label)
        if idx == -1:
            continue
        snippet = text[idx + len(label):]
        # Grab up to 60 chars and parse the first money-like token
        for token in re.split(r"[\s\n]+", snippet[:80]):
            val = parse_money(token)
            if val != 0.0:
                return val
    return 0.0


def _extract_direct_assessments(text: str) -> float:
    """
    Sum all dollar amounts in the 'Direct Assessments' block.
    Stops at the next major section header.
    """
    lower = text.lower()
    start = lower.find("direct assessment")
    if start == -1:
        return 0.0
    # Find end of section (next header keyword)
    end_keywords = ["interest", "total tax", "amount due", "tax summary"]
    end = len(text)
    for kw in end_keywords:
        pos = lower.find(kw, start + 20)
        if pos != -1 and pos < end:
            end = pos

    section = text[start:end]
    amounts = re.findall(r"\$[\d,]+\.?\d*", section)
    return sum(parse_money(a) for a in amounts)


# ── Main loop ─────────────────────────────────────────────────────────────────

async def run(excel_path: str):
    path = Path(excel_path)
    if not path.exists():
        print(f"ERROR: File not found: {excel_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    errors_path = path.parent / "errors.csv"
    errors: list[dict] = []

    total_processed = 0
    total_skipped = 0
    mismatches: list[dict] = []

    print(f"Opening workbook: {excel_path}")
    print(f"Processing rows {FIRST_DATA_ROW} – {LAST_DATA_ROW} (column P)\n")

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=False)  # headed so you can watch
        context = await browser.new_context(viewport={"width": 1400, "height": 900})
        page = await context.new_page()

        for row in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):
            parcel_id = ws.cell(row=row, column=COL_PARCEL).value

            if not parcel_id:
                print(f"  Row {row}: empty parcel ID — skipping")
                total_skipped += 1
                errors.append({"row": row, "parcel_id": "", "reason": "empty parcel ID"})
                continue

            parcel_id = str(parcel_id).strip().zfill(8)
            print(f"  Row {row}: parcel {parcel_id} ...", end=" ", flush=True)

            try:
                data = await scrape_parcel(page, parcel_id)
            except Exception as exc:
                print(f"ERROR ({exc})")
                total_skipped += 1
                errors.append({"row": row, "parcel_id": parcel_id, "reason": str(exc)})
                continue

            if data is None:
                print("not found / no 2024 bill")
                total_skipped += 1
                errors.append({"row": row, "parcel_id": parcel_id, "reason": "no 2024 tax bill found"})
                continue

            # Write to Excel
            ws.cell(row=row, column=COL_ASSESSED).value = data["assessed"]
            ws.cell(row=row, column=COL_MILLAGE).value  = data["millage"]
            ws.cell(row=row, column=COL_DIRECT).value   = data["direct_assessments"]
            ws.cell(row=row, column=COL_INTEREST).value = data["interest"]

            # Validate against column N formula result (if already calculated)
            n_val = ws.cell(row=row, column=COL_FORMULA).value
            if n_val is not None:
                try:
                    n_float = float(n_val)
                    diff = abs(n_float - data["total_tax_bill"])
                    if diff > 1.00:  # allow $1 rounding tolerance
                        mismatches.append({
                            "row": row,
                            "parcel_id": parcel_id,
                            "col_n": round(n_float, 2),
                            "gis_total": round(data["total_tax_bill"], 2),
                            "diff": round(diff, 2),
                        })
                        print(f"OK (⚠ mismatch: N={n_float:.2f}, GIS={data['total_tax_bill']:.2f})")
                    else:
                        print(f"OK ✓ (${data['total_tax_bill']:,.2f})")
                except (TypeError, ValueError):
                    print(f"OK (column N not numeric, skipping validation)")
            else:
                print(f"OK ✓ (${data['total_tax_bill']:,.2f})")

            total_processed += 1

            # Save every N parcels
            if total_processed % SAVE_EVERY == 0:
                wb.save(excel_path)
                print(f"    → Saved progress ({total_processed} parcels done)")

            await asyncio.sleep(DELAY_SECS)

        await browser.close()

    # Final save
    wb.save(excel_path)
    print(f"\n→ Final save complete: {excel_path}")

    # Write errors CSV
    if errors:
        with open(errors_path, "w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=["row", "parcel_id", "reason"])
            writer.writeheader()
            writer.writerows(errors)
        print(f"→ Errors written to: {errors_path}")

    # Summary
    print("\n" + "=" * 55)
    print("SUMMARY")
    print("=" * 55)
    print(f"  Total processed : {total_processed}")
    print(f"  Total skipped   : {total_skipped}")
    print(f"  Mismatches (>$1): {len(mismatches)}")
    if mismatches:
        print("\n  Mismatch details:")
        for m in mismatches:
            print(f"    Row {m['row']} ({m['parcel_id']}): "
                  f"Col N = ${m['col_n']:,.2f}, "
                  f"GIS = ${m['gis_total']:,.2f}, "
                  f"diff = ${m['diff']:,.2f}")
    print("=" * 55)


# ── Entry point ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Mecklenburg County Tax Bill Scraper")
    parser.add_argument(
        "--file",
        required=True,
        help='Path to your Excel file, e.g. --file "C:\\Users\\You\\Desktop\\parcels.xlsx"',
    )
    args = parser.parse_args()
    asyncio.run(run(args.file))
